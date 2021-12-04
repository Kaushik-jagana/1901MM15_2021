# from fpdf.fpdf import _Orientation
# from _typeshed import Self
import streamlit as st
import numpy as np
import pandas as pd
import xlsxwriter 
import openpyxl 
from openpyxl.styles import Border, Font, Alignment, Side
import os
import smtplib 
from email.message import EmailMessage
from fpdf import FPDF


# green_font = Font(name='Century',
#                  size=12,
#                  color='008000')

# red_font = Font(name='Century',
#                  size=12,
#                  color='FF0000')

# blue_font = Font(name='Century',
#                  size=12,
#                  color='0000FF')

# bold_font = Font(name='Century',
#                  bold=True,
#                  size=12)

# thin = Side(border_style="thin")
# border = Border(top=thin, left=thin, right=thin, bottom=thin)

# align = Alignment(horizontal="center")
# st.title("Transcript Generator")

# grades_file = st.file_uploader("upload grades", type=['csv'])

# names_rollno_file = st.file_uploader("upload names-roll", type=['csv'])
# subjects_master_file=st.file_uploader("upload subjects_master",type=['csv'])




def generate_transcript():
    return

class PDF(FPDF):
    


    
pdf= FPDF('L', 'mm', 'A3')
pdf.add_page()
pdf.rect(10,10,400,280)
pdf.line(10,40,410,40)
pdf.line(10,120,410,120)
pdf.line(10,180,410,180)
pdf.line(10,240,410,240)
pdf.image('tittle.jpeg',11,11,399,29)
pdf.rect(100,42,180,18)
pdf.set_font('Arial','B',10)
pdf.set_xy(102,45)
pdf.cell(35,5,'Roll-No')
pdf.cell(35,5,'Name')
pdf.cell(35,5,'Year of Admission')
pdf.set_xy(102,52)
pdf.cell(90,5,'Program:Bachelor of Technology')
pdf.cell(50,5,'Course')

pdf.output('test.pdf','F')

# pdf.output('test.pdf','F')


os.system('test.pdf')
