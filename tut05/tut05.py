
import csv
import os
from openpyxl import Workbook

def get_data():
    dir_name =  "output"
    if not os.path.exists(dir_name):
        os.mkdir(dir_name)
    file1 = open('grades.csv','r')
    csvreader =  csv.reader(file1)
    header = next(csvreader)
    print(header)
    rows = []
    rollnumber = set()
    for row in csvreader:
        rollnumber.add(row[0])
        rows.append(row)
    print(rollnumber)
    #print(rows)
    
    Emptydict = {}
    for row in rows:
        Emptydict[row[0]]=[]
    for row in rows:
        Emptydict[row[0]].append(row)
    print(len(Emptydict[rows[1][0]]),rows[1][0])        
    file2 = open('subjects_master.csv','r')
    csvreader1 = csv.reader(file2)
    header2 = next(csvreader1)
    Dicts2 = {}
    for rows2 in csvreader1:
        Dicts2[rows2[0]]=rows2
    print(Dicts2['CB102'])
    answer = []
    #print(len(Emptydict['0401CS13'][0]),len(Emptydict['0401CS13'][1]),len(Emptydict['0401CS13'][2]))
    for curno in rollnumber:
        gradelist = []
        currlist = []
        sno = 1
        sem = 1
        #f = open("./output/"+curno+".xlsx","w") 
        wb = Workbook()
        
        
        overallworksheet  = wb.create_sheet("Overall")
        sem1sheet = wb.create_sheet("Sem1")
        sem2sheet = wb.create_sheet("Sem2")
        sem3sheet = wb.create_sheet("Sem3")
        sem4sheet = wb.create_sheet("Sem4")
        sem5sheet = wb.create_sheet("Sem5")
        sem6sheet = wb.create_sheet("Sem6")   
        sem7sheet = wb.create_sheet("Sem7")
        sem8sheet = wb.create_sheet("Sem8")
        rows=0     
        sem1sheet['A1'] = 'slno'
        sem1sheet['B1'] = 'subject no'
        sem1sheet['C1'] = 'subject name'
        sem1sheet['D1'] = 'ltp'
        sem1sheet['E1'] = 'credit'
        sem1sheet['F1'] = 'subject type'
        sem1sheet['G1'] = 'grade'
       
        for statusbruh in Emptydict[curno]:
            #print(statusbruh[1])
            if sem != int(statusbruh[1]):
                sem+=1
                sno=1
                currlist.append([])
            tempans = []
            tempans.append(sno)
            tempans.append(statusbruh[2])
            tempans.append(statusbruh[1])
                #print(statusbruh)
            tempans.append(Dicts2[statusbruh[2]][1])
            tempans.append(Dicts2[statusbruh[2]][2])
            tempans.append(Dicts2[statusbruh[2]][3])
            tempans.append(statusbruh[4])
            if(sem==1):
                sem1sheet.append(tempans)
            if(sem==2):
                sem2sheet.append(tempans)
            if(sem==3):
                sem3sheet.append(tempans)
            if(sem==4):
                sem4sheet.append(tempans)
            if(sem==5):
                sem5sheet.append(tempans)
            if(sem==6):
                sem6sheet.append(tempans)
            if(sem==7):
                sem7sheet.append(tempans)
            if(sem==8):
                sem8sheet.append(tempans)  
            sno+=1
            currlist.append(tempans)
        answer.append(currlist)
        wb.save( './output/'+curno + '.xlsx')
        
    #print(answer)
    #print(answer[0])
    #print(Dicts2)
    grades_lines  = file1.readlines()
    #print("Started printing")
    #print(grades_lines)
    #print("bruh")
    i = 0
    for line in grades_lines:
        print(temp)
        temp = line.split(",")
        if i==0:
            grade_names = temp
        else:
            grades.append(temp)
        i += 1
        
    file2 = open('names-roll.csv','r')
    names_lines  = file1.readlines()
    i += 0
    print("Started printing")
    for line in names_lines:
        temp = line.split(",")
        if i==0:
            roll_names = temp
        else:
            name_roll.append(temp)
        i += 1
    
    file3 = open('subjects_master.csv','r')
    grades_lines  = file1.readlines()
    print("Started printing")
    i = 0
    for line in grades_lines:
        temp = line.split(",")
        if i==0:
            subject_names = temp
        else:
            subject_master.append(temp)
        i += 1

def generate_marksheet():
    get_data()



    #print("Finished printing")


grade_names=[]
grades = []
roll_names=[]
name_roll=[]
subject_names = []
subject_master = []
generate_marksheet()
print(grade_names)

