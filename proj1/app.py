import streamlit as st
import numpy as np
import pandas as pd
import xlsxwriter 
import openpyxl 
from openpyxl.styles import Border, Font, Alignment, Side
import os
import smtplib 
from email.message import EmailMessage
import shutil


green_font = Font(name='Century',
                 size=12,
                 color='008000')

red_font = Font(name='Century',
                 size=12,
                 color='FF0000')

blue_font = Font(name='Century',
                 size=12,
                 color='0000FF')

bold_font = Font(name='Century',
                 bold=True,
                 size=12)

thin = Side(border_style="thin")
border = Border(top=thin, left=thin, right=thin, bottom=thin)

align = Alignment(horizontal="center")



st.title("Marksheet Generator")

master_roll_file = st.file_uploader("upload master_roll", type=['csv'])

responses_file = st.file_uploader("upload responses", type=['csv'])

col1, col2 = st.columns(2)
with col1:
	p = st.number_input("Enter marks for correct ans: ")
with col2:
	n = st.number_input("Enter -ve marks for wrong ans: ", help='enter with "-" symbol')


st.text("")

# m = st.markdown("""
# <style>
# div.stButton > button:first-child {
#     background-color: lightgreen; 
# }
# </style>""", unsafe_allow_html=True)  


def excel_creator(output_path,correct_ans):
	workbook = xlsxwriter.Workbook(output_path)
	worksheet = workbook.add_worksheet('quiz')

	worksheet.set_column(0,5,17)
	worksheet.set_row(4,22)
	worksheet.insert_image('A1', 'logo.png',{'x_scale':620/853, 'y_scale':82/126})

	merge_format = workbook.add_format({
                                        'bold': 1, 
                                        'align': 'center',
                                        'underline': 1,
                                        'font_size':16,
                                        'font_name':'Century'})
	right_format = workbook.add_format({
                                        'align':'right',
                                        'font_name':'Century',
                                        'font_size':12 })
	bold_format = workbook.add_format({
                                        'bold':1,
                                        'font_name':'Century',
                                        'font_size':12 })
	header_format = workbook.add_format({
                                        'bold':1,
                                        'align':'center',
                                        'font_size':12,
                                        'font_name':'Century',
                                        'border':1})
	border_format = workbook.add_format({
                                        'border':1
                                        })
	blue_format = workbook.add_format({
                                        'border':1,
                                        'font_name':'Century',
                                        'font_size':12,
                                        'font_color':'#0000FF',
                                        'align':'center'})
	black_format = workbook.add_format({
                                        'border':1,
                                        'font_name':'Century',
                                        'font_size':12, 
                                        'align':'center'})

	worksheet.merge_range('A5:E5', 'Marks Sheet', merge_format)
	worksheet.write('A6','Name:',right_format)
	worksheet.write('A7','Roll Number:',right_format)
	worksheet.write('D6','Exam:',right_format)
	worksheet.write('E6','quiz',bold_format)
	worksheet.write('A9',' ',header_format)
	worksheet.write('B9','Right',header_format)
	worksheet.write('C9','Wrong',header_format)
	worksheet.write('D9','Not Attempt',header_format)
	worksheet.write('D12',' ',header_format)
	worksheet.write('E9','Max',header_format)
	worksheet.write('E10',28,black_format)
	worksheet.write('E11',' ',header_format)
	worksheet.write('A10','No.',header_format)
	worksheet.write('A11','Marking',header_format)
	worksheet.write('A12','Total',header_format)
	worksheet.write('A15','Student Ans',header_format)
	worksheet.write('B15','Correct Ans',header_format)
	worksheet.write('D15','Student Ans',header_format)
	worksheet.write('E15','Correct Ans',header_format)

	for i in range(25):
		worksheet.write('B'+str(16+i),correct_ans[i],blue_format)

	for i in range(len(correct_ans)-25):
		worksheet.write('E'+str(16+i),correct_ans[25+i],blue_format)

	workbook.close()
     

def marksheet_creator(roll_number, responses, master_roll):

	output_path = "marksheets/" + roll_number + ".xlsx"

	correct_ans = list(responses[responses['Roll Number']=='ANSWER'].loc[:,'Unnamed: 7':'Unnamed: 34'].iloc[0])

	given_response = list(responses[responses['Roll Number']==roll_number].loc[:,'Unnamed: 7':'Unnamed: 34'].iloc[0])

	excel_creator(output_path, correct_ans)
	workbook = openpyxl.load_workbook(output_path)
	worksheet = workbook.get_sheet_by_name('quiz')
	worksheet['B6'] =  str(master_roll[master_roll['roll']==roll_number]['name'].iloc[0])
	worksheet['B6'].font = bold_font
	worksheet['B7'] = roll_number
	worksheet['B7'].font = bold_font
	correct = 0
	wrong = 0
	not_attempted = 0
	for j in range(len(given_response)):
		if(pd.isnull(given_response[j])):
			not_attempted += 1
		elif(given_response[j]==correct_ans[j]):
			correct += 1
			if(j<25):
				worksheet['A'+str(16+j)] = correct_ans[j] 
				worksheet['A'+str(16+j)].font = green_font
				worksheet['A'+str(16+j)].alignment = align
				worksheet['A'+str(16+j)].border = border
			else:
				worksheet['D'+str(16+j-25)] = correct_ans[j]
				worksheet['D'+str(16+j-25)].font = green_font
				worksheet['D'+str(16+j-25)].alignment = align
				worksheet['D'+str(16+j-25)].border = border
		else:
			wrong += 1
			if(j<25):
				worksheet['A'+str(16+j)] = given_response[j]
				worksheet['A'+str(16+j)].font = red_font
				worksheet['A'+str(16+j)].alignment = align
				worksheet['A'+str(16+j)].border = border
			else:
				worksheet['D'+str(16+j-25)] = given_response[j]
				worksheet['D'+str(16+j-25)].font = red_font
				worksheet['D'+str(16+j-25)].alignment = align
				worksheet['D'+str(16+j-25)].border = border

	# st.write('correct: {}, wrong: {}, not_attempted: {} '.format(correct,wrong,not_attempted))
	worksheet['B10'] = correct
	worksheet['B10'].font = green_font
	worksheet['B10'].alignment = align
	worksheet['B10'].border = border
	worksheet['B11'] =  p
	worksheet['B11'].font = green_font
	worksheet['B11'].alignment = align
	worksheet['B11'].border = border
	worksheet['B12'] =  correct*p
	worksheet['B12'].font = green_font
	worksheet['B12'].alignment = align
	worksheet['B12'].border = border

	worksheet['C10'] = wrong
	worksheet['C10'].font = red_font
	worksheet['C10'].alignment = align
	worksheet['C10'].border = border
	worksheet['C11'] = n
	worksheet['C11'].font = red_font
	worksheet['C11'].alignment = align
	worksheet['C11'].border = border
	worksheet['C12'] = wrong*n
	worksheet['C12'].font = red_font
	worksheet['C12'].alignment = align
	worksheet['C12'].border = border

	worksheet['D10'] = not_attempted
	worksheet['D10'].font = Font(name='Century',size=12)
	worksheet['D10'].alignment = align 
	worksheet['D10'].border = border
	worksheet['D11'] = 0
	worksheet['D11'].font = Font(name='Century',size=12)
	worksheet['D11'].alignment = align
	worksheet['D11'].border = border

	worksheet['E12'] = str(correct*p+wrong*n)+'/'+str(len(correct_ans)*p)
	worksheet['E12'].font = blue_font
	worksheet['E12'].alignment = align
	worksheet['E12'].border = border

	workbook.save(output_path)

	return [correct,wrong,not_attempted]



if(master_roll_file is not None and responses_file is not None and p!=0):
	master_roll = pd.read_csv(master_roll_file)
	mail_ids = master_roll
	responses = pd.read_csv(responses_file)
	responses['Roll Number'] = responses['Roll Number'].str.upper()
	mail_ids['Email'] = responses['Email address']
	mail_ids['Insti_Email'] = responses['IITP webmail']
	roll_nos = list(master_roll['roll'].iloc[:])
	for i in range(len(roll_nos)):
		if(roll_nos[i] not in list(responses['Roll Number'])):
			df = pd.DataFrame([[ master_roll[master_roll['roll']==roll_nos[i]].iat[0,1],roll_nos[i],'Absent']],columns = ['Name','Roll Number','Score'],index=[i-0.5])
			responses = responses.append(df,ignore_index=False)
			responses = responses.sort_index().reset_index(drop=True)



 
def generate_rollwise():
	if(master_roll_file is not None and responses_file is not None and p!=0):

		roll_nos = list(master_roll['roll'].iloc[:])

		Summary = []
		score = []

		if('ANSWER' in list(responses['Roll Number'])):
			
			for i in range(len(roll_nos)):

				# st.write(roll_nos[i])

				if(pd.isnull(responses[responses['Roll Number']==roll_nos[i]]['Score'].iloc[0])):
					output_path = "marksheets/" + roll_nos[i] + ".xlsx"
					workbook = xlsxwriter.Workbook(output_path)
					worksheet = workbook.add_worksheet('quiz')
					workbook.close()
					Summary.append([])
					score.append('0/'+str(p*28))
					# print('Absent')
				else:
					correct,wrong,not_attempted = marksheet_creator(roll_nos[i],responses, master_roll)
					Summary.append([correct,wrong,not_attempted])
					score.append(str(correct*p+wrong*n)+'/'+str(p*28))

		else:
			print('NO ANSWER')

		st.success('Done')

		return Summary, score 
	else:
		st.error("Some files are missing")
        
        
    
   
# _,c2,_ = st.columns(3)
# with c2:

generate_roll = st.button("Generate rollwise&concise marksheet") 

if(generate_roll):
	gif_runner = st.image("https://assets.website-files.com/5509bda13cd56c1a716d17d1/57be02ffdf9a01fe61eb2ef1_0041_17.gif",width=200)
	 
	Summary, score = generate_rollwise() 
	responses['statusAns'] = Summary
	responses['Score_After_Negative'] = score
	responses.columns  = ['Timestamp', 'Email address', 'Google_Score', 'Name', 
	'IITP webmail','Phone (10 digit only)', 'Roll Number', 'Unnamed: 7', 
	'Unnamed: 8','Unnamed: 9', 'Unnamed: 10', 'Unnamed: 11', 'Unnamed: 12',
	'Unnamed: 13', 'Unnamed: 14', 'Unnamed: 15', 'Unnamed: 16','Unnamed: 17',
	'Unnamed: 18', 'Unnamed: 19', 'Unnamed: 20','Unnamed: 21', 'Unnamed: 22', 
	'Unnamed: 23', 'Unnamed: 24','Unnamed: 25', 'Unnamed: 26', 'Unnamed: 27', 
	'Unnamed: 28','Unnamed: 29', 'Unnamed: 30', 'Unnamed: 31', 'Unnamed: 32',
	'Unnamed: 33', 'Unnamed: 34', 'statusAns', 'Score_After_Negative']
	cols = ['Timestamp', 'Email address', 'Google_Score', 'Name', 
	'IITP webmail','Phone (10 digit only)','Score_After_Negative', 'Roll Number', 
	'Unnamed: 7','Unnamed: 8','Unnamed: 9', 'Unnamed: 10', 'Unnamed: 11', 'Unnamed: 12',
	'Unnamed: 13', 'Unnamed: 14', 'Unnamed: 15', 'Unnamed: 16','Unnamed: 17',
	'Unnamed: 18', 'Unnamed: 19', 'Unnamed: 20','Unnamed: 21', 'Unnamed: 22', 
	'Unnamed: 23', 'Unnamed: 24','Unnamed: 25', 'Unnamed: 26', 'Unnamed: 27', 
	'Unnamed: 28','Unnamed: 29', 'Unnamed: 30', 'Unnamed: 31', 'Unnamed: 32',
	'Unnamed: 33', 'Unnamed: 34', 'statusAns']
	responses = responses[cols]
	responses.to_excel("marksheets/concise_marksheet.xlsx", sheet_name='concise_marksheet', index=False)
	gif_runner.empty()




EMAIL_ADDRESS = 'aravindreddy02kappa@gmail.com'
EMAIL_PASSWORD = "Aravind123"

mail = st.button("Send Emails to all the students")

if(mail):
	msg = EmailMessage()
	msg['Subject'] = 'Marksheet'
	msg['From'] = EMAIL_ADDRESS
	# msg['To'] = 'aravindreddy2260@gmail.com'
	msg.set_content('Check out your marks in quiz')
	# file = "sample.xlsx"

	# with open(file, 'rb') as f:
	# 	file_data = f.read()
	# 	file_name = f.name

	# msg.add_attachment(file_data, maintype='application', subtype='octect-stream', filename='sample.xlsx')
	
	# gif_runner = st.image("https://i.pinimg.com/originals/98/64/9a/98649add72e05e3cc1b8ae0e6f553c8e.gif", width=200)
	# with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
	# 	smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
	# 	smtp.send_message(msg)
	# gif_runner.empty()
	# st.success("Done")


	# st.dataframe(mail_ids)
	file_path = "marksheets/"


	 
	for i in list(mail_ids['roll']):
		msg= EmailMessage()
		msg['Subject']='Marksheet'
		msg['From']=EMAIL_ADDRESS
		msg.set_content('Check out your marks in quiz')
		email_list = list(mail_ids[mail_ids['roll']==i].iloc[0,[2,3]])
		msg['To'] =  ','.join(email_list)
		file = 'marksheets/' + i + '.xlsx'

		with open(file, 'rb') as f:
			file_data = f.read()
			

		msg.add_attachment(file_data, maintype='application', subtype='octect-stream',filename=i+'.xlsx')
		server = smtplib.SMTP_SSL('smtp.gmail.com', 465) 
		server.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
		server.send_message(msg)
		server.quit()
		del msg['To']
		st.success("Email sent to "+ mail_ids[mail_ids['roll']==i].iat[0,1])


shutil.make_archive('total_files', 'zip', "marksheets")

with open("total_files.zip", "rb") as fp:
    btn = st.download_button(
        label="Download all files",
        data=fp,
        file_name="total_files.zip",
        mime="application/zip"
    )


